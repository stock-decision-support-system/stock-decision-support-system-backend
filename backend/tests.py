from django.test import TestCase
import shioaji as sj
import yaml
import pandas as pd
import openpyxl
import datetime
from openpyxl.styles import NamedStyle

with open("config.yaml", "r") as file:
     config = yaml.safe_load(file)

wb = openpyxl.load_workbook('stocks.xlsx')
date_style_name = 'datetimes'
date_style = NamedStyle(name='datetimes', number_format='m"月"d"日"')

class ShioajiTestCase(TestCase):
     def test_virtual_account_login(self):
         api = sj.Shioaji(simulation=True) # 模擬模式
         api.login(
             api_key=config["shioaji"]["api_key"],
             secret_key=config["shioaji"]["secret_key"],
             # contracts_cb=lambda security_type: print(
             # f"{repr(security_type)} fetch done."
             # ), # 取得商品文件Callback
         )
         #print(api.stock_account)# 預設賬號
         print(api.Contracts.Stocks["2890"]) #證券
         """
          order = api.Order(
             price=12,
             quantity=1,
             action=sj.constant.Action.Buy,
             price_type=sj.constant.StockPriceType.LMT,
             order_type=sj.constant.OrderType.ROD,
             order_lot=sj.constant.StockOrderLot.Common,
             account=api.stock_account
          )

         # print(api.Contracts) #商品文件資訊
         # print(api.Contracts.Stocks["2890"]) #證券
         # print(api.Contracts.Futures["TXFA3"]) #期貨
         # print(api.Contracts.Options["TXO18000R3"]) #選擇權
         # print(api.Contracts.Indexs.TSE) #顯示所有指數
         # print(api.Contracts.Indexs.TSE["001"]) # 指數
         """

     def test_tick(self): # 整股 即時行情
         api = sj.Shioaji(simulation=True) # 模擬模式
         api.login(
             api_key=config["shioaji"]["api_key"],
             secret_key=config["shioaji"]["secret_key"],
         )
         api.quote.subscribe(
             api.Contracts.Stocks["2330"],
             quote_type=sj.constant.QuoteType.Tick,
             version=sj.constant.QuoteVersion.v1,
         )
        

     def test_order(self):
         api = sj.Shioaji(simulation=True) # 模擬模式
         api.login(
             api_key=config["shioaji"]["api_key"],
             secret_key=config["shioaji"]["secret_key"],
         )
        
         contract = api.Contracts.Stocks.TSE.TSE2890
         order = api.Order(
             price=17,
             quantity=3,
             action=sj.constant.Action.Buy,
             price_type=sj.constant.StockPriceType.LMT,
             order_type=sj.constant.OrderType.ROD,
             order_lot=sj.constant.StockOrderLot.Common,
             # daytrade_short=False,
             custom_field="test",
             account=api.stock_account
         )
         trade = api.place_order(contract, order)
         print(trade)
        
         # trade = api.update_status(api.stock_account) #更新狀態(成交後)
         # print(trade)

         api.update_status(api.stock_account) # 取得證券委託狀態
         print(api.list_trades())
     def test_balance(self):
         api = sj.Shioaji(simulation=True) # 模擬模式
         api.login(
             api_key=config["shioaji"]["api_key"],
             secret_key=config["shioaji"]["secret_key"],
         )
         print(api.account_balance())#查看銀行餘額

     # def test_old_tick(self): # 整股 歷史行情
     # api = sj.Shioaji(simulation=True) # 模擬模式
     # api.login(
     # api_key=config["shioaji"]["api_key"],
     # secret_key=config["shioaji"]["secret_key"],
     # )
     # ticks = api.ticks(
     # contract=api.Contracts.Stocks["2330"],
     # date="2024-03-18",
     # query_type=sj.constant.TicksQueryType.LastCount,
     # last_cnt=1,
     # )
     # print(ticks.close)

     def test_old_tick(self): # 整股 歷史行情 拿來抓每日收盤價
         api = sj.Shioaji(simulation=True) # 模擬模式
         api.login(
             api_key=config["shioaji"]["api_key"],
             secret_key=config["shioaji"]["secret_key"],
         )
        
         stock_list = ["2330", "2454", "2317", "2382", "2308", "2303", "2891", "2881", "3711", "2412", "2886", "2882" , "1216", "2884", "1303", "3034", "3231", "2357", "2885", "2002", "2892", "2345", "5880", "1301", " 3008", "2379", "3037", "5871", "2301", "2890", "2880", "2207", "2327", "2887", "1101", "2883", "1326" , "6669", "4938", "2395", "3045", "5876", "2603", "2912", "1590", "4904", "2801", "6505", "2408", " 9910"]
         close_prices = {} # 儲存每檔股票的close價格
         today = "2024-03-25"
         # today = datetime.date.today()

         for stock_code in stock_list:
             ticks = api.ticks(
                 contract=api.Contracts.Stocks[stock_code],
                 date=str(today),
                 query_type=sj.constant.TicksQueryType.LastCount,
                 last_cnt=1,
                 )
                 # 將每檔股票的close價格儲存到字典中
             close_prices[stock_code] = ticks.close if hasattr(ticks, 'close') else None

         close_price = [price[0] for price in close_prices.values()]
        
         s1 = wb['price']
         row = 19
         col = 2
         flag = False

         while flag == False :
            if s1.cell(row=row,column=col).value is None:
                flag = True
            else:
                row += 1

         for price in close_price:
              s1.cell(row=row, column=col, value=price)
              col += 1

         date_cell = s1.cell(row=row,column=1)
         date_cell.value=today
         date_cell.style=date_style_name
         wb.save('stocks.xlsx')
''' TODO
     def get_user_input(self):
         while True:
             user_input = input("請輸入INV的值：")
             try:
                 # 嘗試將使用者輸入轉換為浮點數
                 inv_value = float(user_input)
                 return inv_value
             except ValueError:
                 # 如果轉換失敗，提醒使用者輸入的不是有效的數字，並提示重新輸入
                 print("輸入無效，請輸入有效的數字。")
    
     def investment_simulation_BH(self):
         weight = wb['weight(BH)'] # 這個在你的程式碼片段中沒有被使用
         holding = wb['holding(BH)']
         INVchange = wb['INVchange(BH)']
         price = wb['price']

         row = 2
         col = 2
         flag = False

         while flag==False:
             if holding.cell(row=row,column=col).value is None:
                 break
             else :
                 row += 1
        
         # INV = float(INVchange.cell(row=row, column=col).value)
         INV = self.get_user_input()
         INVchange.cell(row=row, column=col).value = INV
         print(INV)

         if type(INVchange.cell(row=row-1,column=col+1)) == str :
             total = 0
         else :
             total = float(INVchange.cell(row=row-1,column=col+1).value)
        
         INVchange.cell(row=row,column=col+1).value = total+INV

         while True:
             p = price.cell(row=row, column=col).value # 在迴圈開始時取得p的值
             if p is None:
                 break # 如果p是None，則結束循環

             p = float(p) # 確保p是一個浮點數
             if INV > 0 :
                 b = (INV * 0.02 / p) + holding.cell(row=row-1,column=col).value
             elif INV == 0 :
                 b = holding.cell(row=row-1,column=col).value
             elif INV < 0 :
                 b = (INV * 0.02 / p) + holding.cell(row=row-1,column=col).value

             holding.cell(row=row, column=col, value=b) # 在holding工作表中設定值
             col += 1 # 移到下一列

         wb.save('stocks.xlsx')
 '''